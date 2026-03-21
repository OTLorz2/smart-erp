"""Export utilities for Excel and CSV"""
from typing import List, Dict, Any, Optional
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pydantic import BaseModel


class ColumnConfig(BaseModel):
    """Column configuration for export"""
    key: str
    title: str
    width: int = 15
    format: Optional[str] = None  # number, currency, date, percent


class ExportHelper:
    """Helper class for data export operations"""

    @staticmethod
    def to_excel(
        data: List[Dict[str, Any]],
        columns: List[ColumnConfig],
        filename: str = "export"
    ) -> bytes:
        """Export data to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = filename

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[chr(64 + col_idx)].width = col.width

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col in enumerate(columns, start=1):
                value = row_data.get(col.key)

                # Format value based on type
                if value is not None:
                    if col.format == "currency" and isinstance(value, (int, float)):
                        value = f"¥{value:,.2f}"
                    elif col.format == "number" and isinstance(value, (int, float)):
                        value = round(value, 2)
                    elif col.format == "percent" and isinstance(value, (int, float)):
                        value = f"{value * 100:.1f}%"

                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def to_csv(
        data: List[Dict[str, Any]],
        columns: List[ColumnConfig]
    ) -> str:
        """Export data to CSV string"""
        if not data:
            return ""

        output = BytesIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([col.title for col in columns])

        # Write data
        for row_data in data:
            row = []
            for col in columns:
                value = row_data.get(col.key)
                if value is not None:
                    if col.format == "currency" and isinstance(value, (int, float)):
                        value = f"¥{value:,.2f}"
                    elif isinstance(value, (int, float)):
                        value = round(value, 2)
                row.append(str(value) if value is not None else "")
            writer.writerow(output := output.getvalue().decode('utf-8') if output.tell() > 0 else "")

        output.seek(0)
        return output.getvalue().decode('utf-8')

    @staticmethod
    def generate_template(columns: List[ColumnConfig]) -> bytes:
        """Generate Excel template with headers only"""
        return ExportHelper.to_excel([], columns, "template")