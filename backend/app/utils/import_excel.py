"""Excel import utilities"""
from typing import List, Dict, Any, Optional, Callable, Set
from io import BytesIO
from openpyxl import load_workbook
from pydantic import BaseModel
from datetime import datetime


class ImportSchema(BaseModel):
    """Schema for import validation"""
    columns: List[str]  # Expected column names
    required_columns: List[str] = []  # Required column names
    unique_columns: List[str] = []  # Columns that must be unique


class ImportRow(BaseModel):
    """Single row of imported data"""
    row_number: int
    data: Dict[str, Any]
    errors: List[str] = []


class ImportResult(BaseModel):
    """Result of import operation"""
    success: bool
    total_rows: int
    valid_rows: int
    invalid_rows: int
    rows: List[ImportRow]
    errors: List[str] = []


class ValidationRule(BaseModel):
    """Validation rule for a field"""
    field: str
    rule_type: str  # required, unique, min, max, min_length, max_length, pattern, choices, custom
    value: Any = None
    message: str = ""


class ValidationError(BaseModel):
    """Single validation error"""
    row: int
    field: str
    message: str


class ValidationResult(BaseModel):
    """Result of validation"""
    valid: bool
    errors: List[ValidationError]
    total_checked: int


class ExcelImporter:
    """Excel file importer"""

    def __init__(self, file_bytes: bytes):
        self.file_bytes = file_bytes

    def parse(self, schema: ImportSchema) -> ImportResult:
        """Parse Excel file and extract data"""
        try:
            wb = load_workbook(BytesIO(self.file_bytes), data_only=True)
            ws = wb.active

            # Get headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Validate required columns
            missing_columns = set(schema.required_columns) - set(headers)
            if missing_columns:
                return ImportResult(
                    success=False,
                    total_rows=0,
                    valid_rows=0,
                    invalid_rows=0,
                    rows=[],
                    errors=[f"Missing required columns: {', '.join(missing_columns)}"]
                )

            # Parse rows
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                data = {}
                for col_idx, header in enumerate(headers):
                    if header is not None:
                        value = row[col_idx] if col_idx < len(row) else None
                        data[header] = value

                row_obj = ImportRow(row_number=row_idx, data=data)
                rows.append(row_obj)

            return ImportResult(
                success=True,
                total_rows=len(rows),
                valid_rows=len(rows),
                invalid_rows=0,
                rows=rows
            )

        except Exception as e:
            return ImportResult(
                success=False,
                total_rows=0,
                valid_rows=0,
                invalid_rows=0,
                rows=[],
                errors=[f"Failed to parse Excel: {str(e)}"]
            )

    def validate(
        self,
        data: List[Dict[str, Any]],
        rules: List[ValidationRule],
        unique_check: Optional[Callable] = None
    ) -> ValidationResult:
        """Validate imported data"""
        errors = []
        seen_values: Dict[str, Set[Any]] = {}

        for row_data in data:
            row_num = row_data.get("row_number", 0)

            for rule in rules:
                field = rule.field
                value = row_data.get("data", {}).get(field)

                # Required check
                if rule.rule_type == "required":
                    if value is None or value == "":
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} is required"
                        ))

                # Unique check
                elif rule.rule_type == "unique":
                    if field not in seen_values:
                        seen_values[field] = set()

                    if value is not None and value in seen_values[field]:
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} value must be unique"
                        ))
                    elif value is not None:
                        seen_values[field].add(value)

                # Min value
                elif rule.rule_type == "min":
                    if value is not None and isinstance(value, (int, float)) and value < rule.value:
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} must be at least {rule.value}"
                        ))

                # Max value
                elif rule.rule_type == "max":
                    if value is not None and isinstance(value, (int, float)) and value > rule.value:
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} must be at most {rule.value}"
                        ))

                # Min length
                elif rule.rule_type == "min_length":
                    if value is not None and isinstance(value, str) and len(value) < rule.value:
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} must be at least {rule.value} characters"
                        ))

                # Max length
                elif rule.rule_type == "max_length":
                    if value is not None and isinstance(value, str) and len(value) > rule.value:
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} must be at most {rule.value} characters"
                        ))

                # Pattern (regex)
                elif rule.rule_type == "pattern":
                    import re
                    if value is not None and isinstance(value, str) and not re.match(rule.value, value):
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} format is invalid"
                        ))

                # Choices
                elif rule.rule_type == "choices":
                    if value is not None and value not in rule.value:
                        errors.append(ValidationError(
                            row=row_num,
                            field=field,
                            message=rule.message or f"{field} must be one of {rule.value}"
                        ))

        return ValidationResult(
            valid=len(errors) == 0,
            errors=errors,
            total_checked=len(data)
        )